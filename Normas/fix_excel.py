import openpyxl
import os
import shutil

file_path = 'temp_copy.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Find header column indices
headers = {cell.value.lower().strip(): idx for idx, cell in enumerate(ws[1]) if cell.value}

mat_col = None
for k, v in headers.items():
    if 'matricula' in k or 'registro' in k or 'id' in k:
        mat_col = v
        break

email_col = None
for k, v in headers.items():
    if 'e-mail' in k or 'email' in k:
        email_col = v
        break

seen_matriculas = set()

for row in range(2, ws.max_row + 1):
    mat_cell = ws.cell(row=row, column=mat_col + 1)
    email_cell = ws.cell(row=row, column=email_col + 1)
    
    # Fix email for 92693
    if str(mat_cell.value).strip() == '92693':
        email_cell.value = 'amanda.monterastelli@g.globo'
    
    # Fix duplicates by prepending '0'
    if mat_cell.value is not None:
        mat_val = str(mat_cell.value).strip()
        if mat_val in ['2301', '4410', '473', '4364', '2688']:
            if mat_val in seen_matriculas:
                mat_cell.value = '0' + mat_val
                mat_cell.data_type = 's'  # Ensure it's stored as string
            else:
                seen_matriculas.add(mat_val)

wb.save(file_path)
print("Excel file temp_copy.xlsx updated successfully.")

try:
    shutil.copy2('temp_copy.xlsx', 'Base de Colaboradores Globo para validação do Agente RIT - Abril 2025.xlsx')
    print("Original file overwritten successfully.")
except Exception as e:
    print("Could not overwrite original file:", e)
